import openpyxl

def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_row)

def getColumnCount(file, sheet_Name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    return (sheet.max_column)

def readExcel(file, sheet_Name, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    return sheet.cell(row=row, column=col).value

def writeExcel(file, sheet_Name, row, col, data):
    flag = False
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_Name)
    sheet.cell(row=row, column=col).value = data
    workbook.save(file)
    flag = True
    return flag


